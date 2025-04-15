import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from scipy.stats import linregress
from datetime import datetime
import tempfile
import io
from PIL import Image
import base64
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

# ---- Constants and Configurations ----
APP_VERSION = "1.0.0"
CURRENT_USER = "gdeschenes2"
LAST_UPDATED = "2025-04-14 19:03:14"

# ---- Data Import Functions ----

def import_tm500(file):
    """
    Import and format TM500 file.
    """
    try:
        df_raw = pd.read_csv(file, delimiter="\t", header=0)
        unit_columns = [col for col in df_raw.columns if "Unit" in col]
        df_cleaned = df_raw.drop(columns=unit_columns)
        value_columns = [col for col in df_cleaned.columns if col.startswith("Value")]
        renamed_columns = {value_columns[i]: f"T{i+1}" for i in range(len(value_columns))}
        df_cleaned = df_cleaned.rename(columns=renamed_columns)
        df_cleaned["Time"] = pd.to_datetime(df_cleaned["Time"], format="%H:%M:%S", errors="coerce")
        df_cleaned["Time"] = (df_cleaned["Time"] - df_cleaned["Time"].iloc[0]).dt.total_seconds()
        temp_columns = [col for col in df_cleaned.columns if col.startswith("T") and col[1:].isdigit()]
        return df_cleaned[["Time"] + temp_columns]
    except Exception as e:
        raise ValueError(f"Error processing TM500 file: {e}")

def import_catman(file):
    """
    Import and format CATMAN Excel file.
    """
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
            temp_file.write(file.getbuffer())
            temp_file_path = temp_file.name
        df_raw = pd.read_excel(temp_file_path, engine="openpyxl")
        df_raw.columns = df_raw.iloc[0]
        df_cleaned = df_raw.iloc[48:].reset_index(drop=True)
        df_cleaned = df_cleaned.loc[:, (df_cleaned != -1000000).any(axis=0)]
        df_cleaned.rename(columns={df_cleaned.columns[0]: "Time"}, inplace=True)
        return df_cleaned
    except Exception as e:
        raise ValueError(f"Error processing CATMAN file: {e}")

def create_excel_report(project_details, test_description, raw_data, summary_table, figures):
    """
    Create an Excel report with all the data and plots.
    """
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        # Combine project details and test description into a single table
        combined_info = {
            "Info": ["Project", "Wheel Type", "Wheel Width Config", "Wheel SKU", 
                     "Test Date", "Load (lbs)", "Speed (kph)", "Duty Time (min)"],
            "Details": [
                project_details["Project"],
                project_details["Wheel Type"],
                project_details["Wheel Width Config"],
                project_details["Wheel SKU"],
                test_description["Date"],
                test_description["Load (lbs)"],
                test_description["Speed (kph)"],
                test_description["Duty Time (min)"]
            ]
        }
        combined_info_df = pd.DataFrame(combined_info)
        
        # Write combined project and test details to a single sheet
        combined_info_df.to_excel(writer, sheet_name='Project & Test Info', index=False)
        
        # Adjust column width for the 'Project & Test Info' sheet
        workbook = writer.book
        worksheet = writer.sheets['Project & Test Info']
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)  # Get the column letter
            for cell in column:
                try:  # Necessary to avoid issues with empty cells
                    max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = max_length + 2  # Add some padding
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Write Raw Data
        raw_data.to_excel(writer, sheet_name='Raw Data', index=False)
        
        # Write Summary Table
        summary_table.to_excel(writer, sheet_name='Summary Table', index=False)
        
        # Adjust column width for the 'Summary Table' sheet
        worksheet = writer.sheets['Summary Table']
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)  # Get the column letter
            for cell in column:
                try:  # Necessary to avoid issues with empty cells
                    max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = max_length + 2  # Add some padding
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Add a 'Plots' sheet with all the figures
        plots_sheet = workbook.create_sheet(title='Plots')
        for i, fig in enumerate(figures, start=1):
            # Convert the plotly figure to an image
            img_bytes = fig.to_image(format="png")
            img = Image(io.BytesIO(img_bytes))
            img.anchor = f'A{i * 20}'  # Adjust the anchor position for each plot
            plots_sheet.add_image(img)

    return buffer

# ---- Data Processing Functions ----

def process_temperature_data(df_cleaned, time_column="Time"):
    """
    Process temperature data for regression analysis.
    """
    temp_columns = [col for col in df_cleaned.columns if col.startswith("T") and col[1:].isdigit()]
    results = []
    processed_data = {}

    time_data = pd.to_numeric(df_cleaned[time_column], errors="coerce")

    for col in temp_columns:
        temp_data = pd.to_numeric(df_cleaned[col], errors="coerce")
        
        # Calculate derivative (temperature rate of change)
        dydx = np.gradient(temp_data, time_data)
        
        # Smooth the derivative using a moving average
        window_size = 5  # Adjust this value to change smoothing
        dydx_smooth = pd.Series(dydx).rolling(window=window_size, center=True).mean()

        idx_max = np.argmax(dydx)
        idx_below_stable = next((idx for idx in range(idx_max + 1, len(dydx))
                               if dydx[idx] < 0.2 and np.all(dydx[idx:] < 0.2)), None)

        regression_line = None
        y_intercept = None
        if idx_below_stable:
            slope, intercept, _, _, _ = linregress(time_data[idx_below_stable:], temp_data[idx_below_stable:])
            regression_line = slope * time_data + intercept
            y_intercept = intercept
            results.append({
                "Channel": col,
                "Intercept (B)": f"{y_intercept:.1f}°C",
                "Max Temp": f"{max(temp_data):.1f}°C"
            })

        processed_data[col] = {
            "time_data": time_data,
            "temp_data": temp_data,
            "regression_line": regression_line,
            "y_intercept": y_intercept,
            "derivative": dydx_smooth
        }

    return results, processed_data

def calculate_ambient_temperature(df_cleaned):
    """
    Calculate the ambient temperature by averaging the first 10 data points of 
    each temperature column and then averaging all those averages together.
    """
    temp_columns = [col for col in df_cleaned.columns if col.startswith("T") and col[1:].isdigit()]
    individual_averages = []

    for col in temp_columns:
        first_10_avg = df_cleaned[col].iloc[:10].mean()
        individual_averages.append(first_10_avg)

    ambient_temperature = np.mean(individual_averages)
    return ambient_temperature

def add_adjusted_column(summary_table, ambient_temperature):
    """
    Add a column named 'B_ADJUSTED_38C' to the summary table, calculated as:
    B + (38 - Ambient Temperature) * 0.6.
    """
    # Extract numeric values from the "Intercept (B)" column and calculate B_ADJUSTED_38C
    summary_table["B_ADJUSTED_38C"] = (
        summary_table["Intercept (B)"]
        .str.extract(r"([-+]?\d*\.\d+|\d+)")[0]  # Extract numeric part from "Intercept (B)"
        .astype(float) + (38 - ambient_temperature) * 0.6
    )
    return summary_table

# ---- Plot Building Functions ----

def build_temperature_plot(processed_data, column_name, show_derivative=False):
    """
    Build a Plotly figure for temperature data and regression.
    """
    fig = go.Figure()

    # Add Temperature Plot
    fig.add_trace(
        go.Scatter(
            x=processed_data["time_data"],
            y=processed_data["temp_data"],
            mode="lines",
            name="Temperature",
            line=dict(color="#FF5733", width=2)
        )
    )

    # Add Linear Regression Line
    if processed_data["regression_line"] is not None:
        fig.add_trace(
            go.Scatter(
                x=processed_data["time_data"],
                y=processed_data["regression_line"],
                mode="lines",
                name="Regression",
                line=dict(color="#33FF57", dash="dash", width=2)
            )
        )
        # Add Marker and Label for Y-intercept
        fig.add_trace(
            go.Scatter(
                x=[0],
                y=[processed_data["y_intercept"]],
                mode="markers+text",
                marker=dict(color="#3498DB", size=10, symbol="diamond"),
                text=[f"B = {processed_data['y_intercept']:.2f}°C"],
                textfont=dict(size=14, color="white"),
                textposition="top right",
                name="Intercept (B)"
            )
        )

    # Add Derivative Plot if enabled
    if show_derivative and "derivative" in processed_data:
        fig.add_trace(
            go.Scatter(
                x=processed_data["time_data"],
                y=processed_data["derivative"],
                mode="lines",
                name="Rate of Change (°C/s)",
                line=dict(color="#FF00FF", width=1.5),  # Magenta color for derivative
                yaxis="y2"  # Use secondary y-axis
            )
        )

    # Update Layout with darker background and vertical grid lines
    layout_dict = {
        "title": dict(
            text=column_name,
            x=0.5,
            xanchor="center",
            font=dict(color="white")
        ),
        "xaxis": dict(
            title="Time [s]",
            tickmode="linear",
            dtick=60,
            gridcolor="rgba(255, 255, 255, 0.2)",
            zeroline=True,
            zerolinecolor="rgba(255, 255, 255, 0.5)",
            range=[0, max(processed_data["time_data"])],
            showgrid=True,
            griddash="solid",
            tickfont=dict(color="white"),
            titlefont=dict(color="white")
        ),
        "yaxis": dict(
            title="Temperature [°C]",
            tickmode="linear",
            dtick=10,
            range=[0, 150],
            gridcolor="rgba(255, 255, 255, 0.2)",
            zeroline=True,
            zerolinecolor="rgba(255, 255, 255, 0.5)",
            tickfont=dict(color="white"),
            titlefont=dict(color="white")
        ),
        "plot_bgcolor": "rgb(50, 50, 50)",
        "paper_bgcolor": "rgb(50, 50, 50)",
        "legend": dict(
            x=1,
            y=0,
            xanchor="right",
            yanchor="bottom",
            font=dict(color="white"),
            bgcolor="rgba(0,0,0,0)"
        )
    }

    # Add secondary y-axis for derivative if enabled
    if show_derivative:
        layout_dict["yaxis2"] = dict(
            title="Rate of Change [°C/s]",
            overlaying="y",
            side="right",
            tickfont=dict(color="#FF00FF"),
            titlefont=dict(color="#FF00FF"),
            gridcolor="rgba(255, 255, 255, 0.1)",
            zeroline=True,
            zerolinecolor="rgba(255, 255, 255, 0.5)"
        )

    fig.update_layout(layout_dict)
    return fig

# ---- Streamlit App ----

# Configure Streamlit page
st.set_page_config(layout="wide", page_title="Temperature Analysis GUI")

# Center Page Title
st.markdown("<h1 style='text-align: center;'>📈 Temperature Analysis GUI</h1>", unsafe_allow_html=True)

# Add version info and last updated timestamp in a small text below the title
st.markdown(f"""
<div style='text-align: center; font-size: 0.8em; color: gray;'>
    Version {APP_VERSION} | Last Updated: {LAST_UPDATED} | User: {CURRENT_USER}
</div>
""", unsafe_allow_html=True)

# Settings Section
st.markdown("## ⚙️ Settings")
col1, col2 = st.columns([1, 2])
with col1:
    doc_type = st.selectbox(
        "Select Document Type:",
        options=["TM500", "CATMAN"],
        index=0,
        key="doc_type_selection"
    )
with col2:
    file_type = ["xls"] if doc_type == "TM500" else ["xlsx"]
    uploaded_file = st.file_uploader(f"Upload your file ({', '.join(file_type).upper()} only)", type=file_type, key="file_upload")

# Project Details Section
st.markdown("### 🚀 Project Details")
col1, col2, col3, col4 = st.columns(4)
with col1:
    project_name = st.text_input("Project:", placeholder="Enter project name", key="project_name")
with col2:
    wheel_type = st.text_input("Wheel Type:", placeholder="Enter wheel type (e.g., IDLER, MIDROLLER)", key="wheel_type_input")
with col3:
    wheel_width_config = st.text_input("Wheel Width Config:", placeholder="Enter width config", key="wheel_width_config")
with col4:
    wheel_sku = st.text_input("Wheel SKU:", placeholder="Enter SKU", key="wheel_sku")

# Test Description Section
st.markdown("### 📝 Test Description")
col1, col2, col3, col4 = st.columns(4)
with col1:
    test_date = st.date_input("Date:", value=datetime.strptime(LAST_UPDATED, "%Y-%m-%d %H:%M:%S").date(), key="test_date")
with col2:
    load = st.text_input("Load [lbs]:", placeholder="Enter load in lbs", key="load_input")
with col3:
    speed = st.text_input("Speed [kph]:", placeholder="Enter speed in kph", key="speed_input")
with col4:
    duty_time = st.text_input("Duty Time [min]:", placeholder="Enter duty time in minutes", key="duty_time_input")

st.markdown("---")

# Main Analysis Section
if uploaded_file:
    try:
        # Import Data
        if doc_type == "TM500":
            df_cleaned = import_tm500(uploaded_file)
        elif doc_type == "CATMAN":
            df_cleaned = import_catman(uploaded_file)
        else:
            st.error("Unsupported document type.")
            st.stop()

        # Display Raw Data (Collapsed by default)
        with st.expander("📋 Raw Data Preview", expanded=False):
            st.dataframe(df_cleaned.head(20))

        # Process Data
        results, processed_data = process_temperature_data(df_cleaned)
        
        # Calculate Ambient Temperature
        ambient_temperature = calculate_ambient_temperature(df_cleaned)

        # Add Adjusted Column to Summary Table
        results_df = pd.DataFrame(results)
        summary_table = add_adjusted_column(results_df, ambient_temperature)

        # Temperature Plots Section with Derivative Toggle
        st.markdown("### 📊 Temperature Plots")
        show_derivative = st.checkbox("Show Temperature Rate of Change", value=False, key="show_derivative")
        
        # Display Plots
        figures = []
        columns = st.columns(3)
        for i, col in enumerate(processed_data.keys()):
            with columns[i % 3]:
                fig = build_temperature_plot(processed_data[col], col, show_derivative)
                st.plotly_chart(fig, use_container_width=True)
                figures.append(fig)

        # Display Ambient Temperature
        st.markdown(f"**Ambient Temperature:** {ambient_temperature:.2f} °C")

        # Display Summary Table
        st.markdown("### Summary Table")
        st.table(summary_table)

        # Add Export Section with prominent button
        st.markdown("### 📥 Export Data")
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            if st.button("Export to Excel", key="export_button", use_container_width=True):
                # Prepare data for export
                project_details = {
                    "Project": project_name,
                    "Wheel Type": wheel_type,
                    "Wheel Width Config": wheel_width_config,
                    "Wheel SKU": wheel_sku
                }
                
                test_description = {
                    "Date": test_date.strftime("%Y-%m-%d"),
                    "Load (lbs)": load,
                    "Speed (kph)": speed,
                    "Duty Time (min)": duty_time
                }
                
                try:
                    # Generate Excel file
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    filename = f"temperature_analysis_{timestamp}.xlsx"
                    
                    # Create Excel report
                    excel_buffer = create_excel_report(
                        project_details,
                        test_description,
                        df_cleaned,
                        summary_table,
                        figures
                    )
                    
                    # Convert to base64 and create download link
                    b64 = base64.b64encode(excel_buffer.getvalue()).decode()
                    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}">📥 Click here to download Excel report</a>'
                    st.markdown(href, unsafe_allow_html=True)
                    
                except Exception as e:
                    st.error(f"Error generating Excel report: {str(e)}")

    except Exception as e:
        st.error(f"An error occurred: {e}")
else:
    st.warning("Please upload a file to proceed.")